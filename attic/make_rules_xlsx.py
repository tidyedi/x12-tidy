"""Generate docs/rules.xlsx.

Sheet 1 "Standard gate"  — locked. Fast path: locate the ISA line, prove it is
                          byte-standard, then parse it at fixed offsets.
Sheet 2 "Recovery draft" — not yet worked. Creative parsing when the gate fails.

Vars:  dirty    = raw file bytes as received.
       cleansed = dirty with any bytes before the ISA tag removed.

Regenerate:  uv run --with openpyxl python make_rules_xlsx.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "docs" / "rules.xlsx"

TOP = Alignment(wrap_text=True, vertical="top")
HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GREEN = PatternFill("solid", fgColor="E7F0E4")

GATE_HEADERS = ["Step", "Statement", "Type"]
GATE_ROWS = [
    (1, "isa_start = dirty.find('ISA')", "assign"),
    (2, "isa_in_document = isa_start != -1", "assign"),
    (3, "if not isa_in_document  ->  identify and exit", "guard"),
    (4, "cleansed = dirty[isa_start:]", "assign"),
    (5, "isa_in_first_position = isa_start == 0", "assign"),
    (6, "too_short_for_isa_line = len(cleansed) < 109", "assign"),
    (7, "if not isa_in_first_position  ->  identify", "guard"),
    (8, "if too_short_for_isa_line  ->  identify and exit", "guard"),
    (9, "element_separator = cleansed[3]", "assign"),
    (10, "gs_in_correct_position = cleansed[106:109] == 'GS' + element_separator", "assign"),
    (11, "if not gs_in_correct_position  ->  recovery", "guard"),
    (12, "else  ->  parse the ISA line", "guard"),
]
GATE_NOTES = [
    "dirty = raw file bytes as received. cleansed = dirty with any bytes "
    "before the ISA tag removed.",
    '"identify and exit" = emit a diagnostic and stop (fatal). '
    '"identify" = emit a diagnostic and keep going.',
    "Parse the ISA line: line = cleansed[0:105]; split on element_separator -> "
    "ISA01..ISA16; component element separator = cleansed[104]; segment "
    "terminator = cleansed[105].",
    "STATUS: standard gate drafted, not yet reviewed rule-by-rule. Recovery "
    "path started on the 'Recovery' sheet.",
]

REC_HEADERS = ["Step", "Statement", "Type"]
REC_ROWS = [
    ("", "Entered from the gate when gs_in_correct_position is false.", "note"),
    (1, "split cleansed on the element separator", "step"),
    (2, "from the 16th element, look for the first 'GS' + element separator",
        "step"),
]
REC_NOTES = [
    "STATUS: recovery path -- only the two steps the user has specified so far. "
    "Build out from here step by step.",
]


def write_sheet(ws, headers, rows, widths, green_cols=()):
    ws.append(headers)
    for col, w in enumerate(widths, start=1):
        c = ws.cell(row=1, column=col)
        c.fill, c.font, c.alignment = HEADER_FILL, HEADER_FONT, TOP
        ws.column_dimensions[get_column_letter(col)].width = w
    for row in rows:
        ws.append(list(row))
        for idx, cell in enumerate(ws[ws.max_row], start=1):
            cell.alignment = TOP
            if idx in green_cols:
                cell.fill = GREEN
    ws.freeze_panes = "A2"


def main() -> None:
    wb = Workbook()

    gate = wb.active
    gate.title = "Standard gate"
    write_sheet(gate, GATE_HEADERS, GATE_ROWS, [6, 72, 10])
    gate.append([])
    for note in GATE_NOTES:
        gate.append(["", note])
        gate.cell(row=gate.max_row, column=2).alignment = TOP

    rec = wb.create_sheet("Recovery")
    write_sheet(rec, REC_HEADERS, REC_ROWS, [6, 78, 10])
    rec.append([])
    for note in REC_NOTES:
        rec.append(["", note])
        rec.cell(row=rec.max_row, column=2).alignment = TOP

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
